#!/usr/bin/env python3 -u

import argparse
from typing import List
from openpyxl.worksheet.worksheet import Worksheet
import schedule
import time
import multiprocessing as mp
from scrapy.crawler import CrawlerProcess, Crawler
from scrapy import signals
from scrapy.utils.project import get_project_settings
from conf import INTERVAL_MINUTES, SAVE_ON_SCRAPE
from njuskalo_scraper.spiders.njuskalo_spider import NjuskaloSpider
from njuskalo_scraper.database import init_database, NjuskaloAdDB
from njuskalo_scraper.util import Notifier, parse_urls_file
from win10toast import ToastNotifier
from openpyxl import Workbook
import datetime
import os

notifier = Notifier()


def _crawl(queue, urls):
    print(datetime.datetime.now(), "Crawling on url", urls)
    items_scraped = []

    def item_scraped(item, response, spider):
        items_scraped.append(item)

    process = CrawlerProcess(get_project_settings())
    crawler = Crawler(NjuskaloSpider, get_project_settings())
    crawler.signals.connect(item_scraped, signals.item_scraped)
    process.crawl(crawler, start_urls=urls)
    process.start()
    queue.put(items_scraped)


def crawl_njuskalo(urls=None):
    queue = mp.Queue()
    process = mp.Process(target=_crawl, args=(queue, urls))
    process.start()
    items_scraped = queue.get()
    process.join()

    print(datetime.datetime.now(), "Scrapping finished")
    if items_scraped:
        lastInterval = datetime.datetime.now() - datetime.timedelta(seconds=INTERVAL_MINUTES * 60 + 5) 
        items = NjuskaloAdDB.select().where(NjuskaloAdDB.scrappedDate >= lastInterval)
        notifier.new_items_received(items)
        saveAllData(items)

def saveAllData(items):
    if(SAVE_ON_SCRAPE and len(items) > 0): 
        print(datetime.datetime.now(), "Started saving data")
        date = datetime.datetime.now().strftime("%Y-%m-%dT%H-%M-%S")
        isExist = os.path.exists("reports")
        if not isExist:
            os.makedirs("reports")
        
        extractToExcel(os.path.join("reports", f"{date} - {len(items)} items.xlsx"), items)

def itemToSheet(ws: Worksheet, row, item):
    col = 1
    for val in item:
        ws.cell(row=row, column=col).value = val
        col += 1
    return ws


def itemsToSheet(ws: Worksheet, row, items):
    for item in items:
        data = vars(item)['__data__'].values()

        col = 1
        for val in data:
            ws.cell(row=row, column=col).value = val
            col += 1

        row += 1
    return ws


def extractToExcel(fileName, items=None):
    items = items if items else NjuskaloAdDB.select()
    print(datetime.datetime.now(), f"Saving {len(items)} items")
    if(len(items) == 0):
        print(datetime.datetime.now(), "No items in database")
        return

    workbook = Workbook()
    njuskalo = workbook['Sheet']
    njuskalo.title = "Njuskalo"

    headers = vars(items[0])['__data__'].keys()
    njuskalo = itemToSheet(njuskalo, 1, headers)
    njuskalo = itemsToSheet(njuskalo, 2, items)
    workbook.save(filename=fileName)


def parse_input():
    args = argparse.ArgumentParser()
    group = args.add_mutually_exclusive_group(required=True)
    group.add_argument('-f',  help='Path to file with urls, one url per line')
    group.add_argument('-e',  help='Path for extracting data')
    args = args.parse_args()
    return args


def start_crawl(urls):
    crawl_njuskalo(urls)
    schedule.every(INTERVAL_MINUTES).minutes.do(crawl_njuskalo, urls=urls)
    while True:
        try:
            schedule.run_pending()
            time.sleep(1)
        except KeyboardInterrupt:
            break


def main():
    init_database()
    input = parse_input()
    try:
        if(input.f):
            urls = parse_urls_file(input.f)
            print(datetime.datetime.now(),f"Scrapping on {urls}")
            start_crawl(urls)
        elif (input.e):
            extractToExcel(input.e)
        else:
            raise "Use -e fileName.xlsx for extracting OR -f fileName.txt for scrapping"
    finally:
        print(datetime.datetime.now(), "Shutting down")


if __name__ == '__main__':
    main()
